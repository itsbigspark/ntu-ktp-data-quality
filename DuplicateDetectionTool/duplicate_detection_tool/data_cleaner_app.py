import streamlit as st
import pandas as pd
from sklearn.metrics import classification_report
from datacleaner.error_detection import ErrorDetector
from datacleaner.error_correction import ErrorCorrector
from datacleaner.utils import detect_column_types
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO

def calculate_metrics(error_data, corrected_data, reference_data):
    """
    Calculate precision, recall, F1-score, and accuracy for error detection and correction.
    """
    reference_data = reference_data.reindex_like(error_data)
    y_true = (error_data != reference_data).any(axis=1).astype(int)
    y_pred = (error_data != corrected_data).any(axis=1).astype(int)
    return classification_report(y_true, y_pred, output_dict=True)

def generate_highlighted_excel(error_data, corrected_data):
    """
    Generate Excel files highlighting errors in RED and corrections in GREEN.
    """
    # Error-highlighted file
    error_workbook = Workbook()
    error_sheet = error_workbook.active
    error_sheet.title = "Error Highlighted"
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for i, col in enumerate(error_data.columns, start=1):
        error_sheet.cell(1, i, col)  # Write header
        for j, value in enumerate(error_data[col], start=2):
            cell = error_sheet.cell(j, i, value)
            if error_data.iloc[j - 2, i - 1] != corrected_data.iloc[j - 2, i - 1]:
                cell.fill = red_fill  # Highlight cell in RED if error detected

    error_file = BytesIO()
    error_workbook.save(error_file)
    error_file.seek(0)

    # Correction-highlighted file
    corrected_workbook = Workbook()
    corrected_sheet = corrected_workbook.active
    corrected_sheet.title = "Corrected Highlighted"
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for i, col in enumerate(corrected_data.columns, start=1):
        corrected_sheet.cell(1, i, col)  # Write header
        for j, value in enumerate(corrected_data[col], start=2):
            cell = corrected_sheet.cell(j, i, value)
            if error_data.iloc[j - 2, i - 1] != corrected_data.iloc[j - 2, i - 1]:
                cell.fill = green_fill  # Highlight cell in GREEN if corrected

    corrected_file = BytesIO()
    corrected_workbook.save(corrected_file)
    corrected_file.seek(0)

    return error_file, corrected_file

# Streamlit App
st.title("🛠️ Error Detection & Correction Tool")
st.write("Upload datasets to detect and correct errors automatically.")

# Add Company Images
#st.image(["images/bigspark_logo.png", "images/UKRI_logo.png", "images/NTU_Primary_logo.png" ], width=200, caption=["Company 1", "Company 2", "Company 3"])

# Add Company Images with Spacing
#st.write("### Partners & Sponsors")
col1, col2, col3 = st.columns([1, 1, 1])  # Three equal-width columns

with col1:
    st.image("images/bigspark_logo.png", width=200)
with col2:
    st.image("images/UKRI_logo.png", width=300)
with col3:
    st.image("images/NTU_Primary_logo.png", width=200)


# File Upload Section
error_file = st.file_uploader("Upload Error Data File (.csv)", type="csv")
reference_file = st.file_uploader("Upload Reference Data File (.csv) (Optional)", type="csv")

if error_file:
    error_data = pd.read_csv(error_file)
    reference_data = pd.read_csv(reference_file) if reference_file else error_data.copy()

    st.subheader("Uploaded Data Previews")
    st.write("Error Data")
    st.dataframe(error_data.head())
    if reference_file:
        st.write("Reference Data")
        st.dataframe(reference_data.head())

    # Detect and Correct Errors
    st.subheader("Error Detection and Correction")
    numeric_features, categorical_features = detect_column_types(error_data)
    corrector = ErrorCorrector(reference_data)
    corrected_data = error_data.copy()

    corrected_data = corrector.correct_missing_values(corrected_data)
    for col in categorical_features:
        corrected_data = corrector.correct_typographical_errors(corrected_data, col)
    for col in numeric_features:
        corrected_data = corrector.correct_out_of_range(corrected_data, col)

    # Display Metrics
    st.subheader("Metrics")
    metrics = calculate_metrics(error_data, corrected_data, reference_data)
    st.json(metrics)
    st.bar_chart(pd.DataFrame(metrics).T[["precision", "recall", "f1-score"]])

    # Display Corrected Data
    st.subheader("Corrected Data Preview")
    st.dataframe(corrected_data.head())

    # Generate Highlighted Excel Files
    error_highlighted, corrected_highlighted = generate_highlighted_excel(error_data, corrected_data)

    # Download Corrected Dataset
    corrected_csv = corrected_data.to_csv(index=False).encode("utf-8")
    st.download_button(
        "Download Corrected Data (CSV)",
        corrected_csv,
        "corrected_data.csv",
        "text/csv",
        key="download-corrected",
    )

    # Download Error Highlighted File
    st.download_button(
        "Download Error Highlighted File (Excel)",
        error_highlighted,
        "error_highlighted.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download-error-highlighted",
    )

    # Download Correction Highlighted File
    st.download_button(
        "Download Correction Highlighted File (Excel)",
        corrected_highlighted,
        "corrected_highlighted.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download-corrected-highlighted",
    )
