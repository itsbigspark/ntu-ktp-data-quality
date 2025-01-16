import pandas as pd
import sys
import json
import argparse
import os
from data_validator.patterns import infer_regex_patterns
from data_validator.cleaning import create_cleaning_rules, apply_cleaning_rules
from data_validator.metrics import calculate_metrics, validate_patterns
from data_validator.visualization import visualize_issues, visualize_metrics
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description="Data Validator CLI")
    parser.add_argument("clean_data", help="Path to clean dataset (CSV)")
    parser.add_argument("unclean_data", help="Path to unclean dataset (CSV)")
    parser.add_argument("--rules", help="Path to custom rules JSON file (optional)", default=None)
    args = parser.parse_args()

    # Ensure output directories exist
    os.makedirs("outputs", exist_ok=True)
    os.makedirs("outputs/visualizations", exist_ok=True)

    # Load datasets
    print("Loading datasets...")
    try:
        clean_data = pd.read_csv(args.clean_data, dtype=str, na_filter=False)
        unclean_data = pd.read_csv(args.unclean_data, dtype=str, na_filter=False)
    except Exception as e:
        print(f"Error loading datasets: {e}")
        sys.exit(1)

    # Initialize variables for patterns and metrics
    patterns = {}
    metrics = {}

    # Check if custom rules are provided
    if args.rules:
        print(f"Using custom rules from: {args.rules}")
        try:
            with open(args.rules, "r") as f:
                custom_rules = json.load(f)
            
            # Extract patterns and metrics if available
            patterns = custom_rules.get("patterns", {})
            metrics = custom_rules.get("metrics", {})
        except Exception as e:
            print(f"Error reading custom rules JSON: {e}")
            sys.exit(1)

    # Infer missing components (patterns or metrics) from the clean dataset
    if not patterns:
        print("No regex patterns provided in JSON. Inferring patterns from clean dataset...")
        patterns = infer_regex_patterns(clean_data)
    if not metrics:
        print("No AVH metrics provided in JSON. Calculating metrics from clean dataset...")
        metrics = calculate_metrics(clean_data)

    # Validate patterns
    try:
        print("Validating patterns...")
        for col, pattern in patterns.items():
            if not isinstance(pattern, str) and not hasattr(pattern, "pattern"):
                raise ValueError(f"Invalid regex pattern for column: {col}")
    except Exception as e:
        print(f"Error in pattern validation: {e}")
        sys.exit(1)

    # Generate cleaning rules based on patterns
    print("Generating cleaning rules...")
    cleaning_rules = create_cleaning_rules(patterns, metrics)

    # Validate and clean the unclean dataset
    print("Validating and cleaning the unclean dataset...")
    cleaned_data, highlighted_issues = apply_cleaning_rules(unclean_data, cleaning_rules)

    # Save outputs
    print("Saving outputs...")
    try:
        # Save cleaned data
        cleaned_data.to_csv("outputs/cleaned_data.csv", index=False)
        print("- Cleaned data saved as: outputs/cleaned_data.csv")

        # Save highlighted issues with formatting
        save_and_highlight_excel(highlighted_issues, "outputs/highlighted_issues.xlsx")
        print("- Highlighted issues saved as: outputs/highlighted_issues.xlsx")

        # Save patterns and metrics
        with open("outputs/metrics.json", "w") as f:
            json.dump({"patterns": patterns, "metrics": metrics}, f)
        print("- Metrics and patterns saved as: outputs/metrics.json")

        # Visualize issues and metrics
        if not highlighted_issues.empty:
            visualize_issues(highlighted_issues, "outputs/visualizations/issue_distribution.png")
        else:
            print("No issues to visualize.")

        if metrics:
            visualize_metrics(metrics, "outputs/visualizations/metrics_plot.png")

        print("- Visualizations saved in: outputs/visualizations/")
        print("\nProcessing complete. Outputs saved in 'outputs/' directory.")
    except Exception as e:
        print(f"Error saving outputs: {e}")
        sys.exit(1)




def save_and_highlight_excel(dataframe, file_path):
    """
    Save a dataframe to Excel and apply conditional formatting:
    - Yellow fill for 'MISSING'
    - Red fill for 'INVALID'
    """
    try:
        # Save the dataframe to an Excel file
        dataframe.to_excel(file_path, index=False, engine="openpyxl")
        
        # Open the workbook and get the active sheet
        workbook = load_workbook(file_path)
        worksheet = workbook.active

        # Define the styles for highlighting
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Apply styles to cells based on their value
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if cell.value == "MISSING":  # Check for MISSING
                    cell.fill = yellow_fill
                elif cell.value == "INVALID":  # Check for INVALID
                    cell.fill = red_fill

        # Save the workbook with applied formatting
        workbook.save(file_path)
        print(f"Conditional formatting applied and saved to {file_path}")
    except Exception as e:
        print(f"Error applying conditional formatting: {e}")



if __name__ == "__main__":
    main()
