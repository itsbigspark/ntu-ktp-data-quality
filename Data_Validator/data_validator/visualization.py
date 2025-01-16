import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import pandas as pd


def color_code_highlighted_issues(file_path):
    """
    Open the highlighted issues Excel file and apply color coding:
    - 'INVALID' cells with green.
    - 'MISSING' cells with red.
    
    Parameters:
        file_path (str): Path to the Excel file with highlighted issues.
    """
    try:
        print(f"Applying color coding to {file_path}...")
        wb = load_workbook(file_path)
        ws = wb.active

        # Define color fills
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        for row in ws.iter_rows(min_row=2):  # Skip the header row
            for cell in row:
                if cell.value == "MISSING":
                    cell.fill = red_fill
                elif cell.value == "INVALID":
                    cell.fill = green_fill

        wb.save(file_path)
        print(f"Color-coded issues successfully saved to {file_path}")
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}.")
    except Exception as e:
        print(f"Error applying color coding: {e}")


def visualize_metrics(history, output_path="outputs/visualizations/metrics_plot.png"):
    """
    Visualize impurity and FPR over iterations and save the plot.
    
    Parameters:
        history (list of dict): A list of dictionaries containing 'impurity' and 'fpr' values.
        output_path (str): Path to save the visualization.
    """
    try:
        if not history or not isinstance(history, list):
            print("Invalid or empty history data provided for metrics visualization.")
            return

        impurity = [h.get("impurity", 0) for h in history]
        fpr = [h.get("fpr", 0) for h in history]

        plt.figure(figsize=(10, 6))
        plt.plot(impurity, label="Impurity", marker="o", color="blue")
        plt.plot(fpr, label="FPR", marker="o", color="orange")
        plt.xlabel("Iterations")
        plt.ylabel("Metrics")
        plt.title("Impurity and FPR over Iterations")
        plt.legend()
        plt.grid(True)

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        plt.savefig(output_path)
        plt.close()
        print(f"Metrics visualization saved to {output_path}")
    except Exception as e:
        print(f"Error generating metrics visualization: {e}")


def visualize_issues(highlighted_data, output_path="outputs/visualizations/issue_distribution.png"):
    """
    Visualize the distribution of 'MISSING' and 'INVALID' issues in the dataset.
    
    Parameters:
        highlighted_data (pd.DataFrame): The dataset with issues highlighted as 'MISSING' or 'INVALID'.
        output_path (str): Path to save the visualization.
    """
    try:
        if not isinstance(highlighted_data, pd.DataFrame) or highlighted_data.empty:
            print("Invalid or empty data provided for issue distribution visualization.")
            return

        # Calculate issue counts for columns
        issue_counts = highlighted_data.isin(["MISSING", "INVALID"]).sum()

        # Create bar plot
        plt.figure(figsize=(12, 6))
        issue_counts.sort_values().plot(kind="bar", color="skyblue")
        plt.title("Distribution of Highlighted Issues (MISSING/INVALID) by Column")
        plt.xlabel("Columns")
        plt.ylabel("Count of Issues")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        plt.savefig(output_path)
        plt.close()
        print(f"Issue distribution visualization saved to {output_path}")
    except Exception as e:
        print(f"Error generating issue distribution visualization: {e}")


def visualize_avh_constraints(metrics, output_path="outputs/visualizations/avh_constraints.png"):
    """
    Visualize AVH metrics (completeness ratio, entropy, and unique ratio) across dataset columns.

    Parameters:
        metrics (dict): Dictionary containing AVH metrics for each column.
        output_path (str): Path to save the visualization.
    """
    try:
        if not metrics or not isinstance(metrics, dict):
            print("Invalid or empty metrics data provided for AVH visualization.")
            return

        # Extract metrics
        columns = list(metrics.keys())
        completeness_ratio = [v.get("completeness_ratio", 0) for v in metrics.values()]
        entropy = [v.get("entropy", 0) for v in metrics.values()]
        unique_ratio = [v.get("unique_ratio", 0) for v in metrics.values()]

        # Create bar plot for AVH metrics
        x = range(len(columns))
        width = 0.25

        plt.figure(figsize=(14, 8))
        plt.bar(x, completeness_ratio, width=width, label="Completeness Ratio", color="green", alpha=0.7)
        plt.bar([i + width for i in x], entropy, width=width, label="Entropy", color="blue", alpha=0.7)
        plt.bar([i + 2 * width for i in x], unique_ratio, width=width, label="Unique Ratio", color="orange", alpha=0.7)

        plt.xticks([i + width for i in x], columns, rotation=45, ha="right")
        plt.xlabel("Columns")
        plt.ylabel("Metrics")
        plt.title("AVH Metrics by Column")
        plt.legend()
        plt.tight_layout()

        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        plt.savefig(output_path)
        plt.close()
        print(f"AVH metrics visualization saved to {output_path}")
    except Exception as e:
        print(f"Error generating AVH metrics visualization: {e}")


def format_highlighted_issues(file_path, save_path=None):
    """
    Open the highlighted issues CSV, convert it to Excel, and apply formatting.

    Parameters:
        file_path (str): Path to the highlighted issues CSV file.
        save_path (str): Path to save the formatted Excel file (optional).
    """
    try:
        # Read the CSV file into a DataFrame
        data = pd.read_csv(file_path)

        # Define save path
        if save_path is None:
            save_path = file_path.replace(".csv", ".xlsx")

        # Write to Excel
        data.to_excel(save_path, index=False)

        # Apply color coding
        color_code_highlighted_issues(save_path)

        print(f"Formatted highlighted issues saved to {save_path}")
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}.")
    except Exception as e:
        print(f"Error formatting highlighted issues: {e}")
