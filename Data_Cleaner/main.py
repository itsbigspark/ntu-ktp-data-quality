import sys
import pandas as pd
from sklearn.metrics import classification_report
from datacleaner.error_detection import ErrorDetector
from datacleaner.error_correction import ErrorCorrector
from datacleaner.utils import detect_column_types
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def calculate_metrics(error_data, corrected_data, reference_data):
    """
    Calculate precision, recall, F1-score, and accuracy for error detection and correction.
    """
    print("Calculating metrics...")

    # Ensure alignment of reference_data with error_data
    reference_data = reference_data.reindex_like(error_data)

    # Identify true errors and detected/corrected errors
    y_true = (error_data != reference_data).any(axis=1).astype(int)  # True error locations
    y_pred = (error_data != corrected_data).any(axis=1).astype(int)  # Detected/corrected locations

    # Generate classification report
    report = classification_report(y_true, y_pred, target_names=["Correct", "Error"])
    #print("\nError Detection Metrics:")
    #print(report)
    #return report

def highlight_errors_and_corrections(error_data, corrected_data, output_error_file, output_corrected_file):
    """
    Create two Excel files:
    - One highlighting errors in RED.
    - One highlighting corrections in GREEN.
    """
    print("Generating error-highlighted file...")
    error_workbook = Workbook()
    error_sheet = error_workbook.active
    error_sheet.title = "Error Highlighted"

    # Highlight errors in RED
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    for i, col in enumerate(error_data.columns, start=1):
        error_sheet.cell(1, i, col)  # Write header
        for j, value in enumerate(error_data[col], start=2):
            cell = error_sheet.cell(j, i, value)
            if error_data.iloc[j - 2, i - 1] != corrected_data.iloc[j - 2, i - 1]:
                cell.fill = red_fill  # Highlight cell in RED if error detected

    error_workbook.save(output_error_file)
    print(f"Error-highlighted file saved to {output_error_file}.")

    print("Generating corrected-highlighted file...")
    corrected_workbook = Workbook()
    corrected_sheet = corrected_workbook.active
    corrected_sheet.title = "Corrected Highlighted"

    # Highlight corrections in GREEN
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    for i, col in enumerate(corrected_data.columns, start=1):
        corrected_sheet.cell(1, i, col)  # Write header
        for j, value in enumerate(corrected_data[col], start=2):
            cell = corrected_sheet.cell(j, i, value)
            if error_data.iloc[j - 2, i - 1] != corrected_data.iloc[j - 2, i - 1]:
                cell.fill = green_fill  # Highlight cell in GREEN if corrected

    corrected_workbook.save(output_corrected_file)
    print(f"Corrected-highlighted file saved to {output_corrected_file}.")

def main():
    if len(sys.argv) < 2:
        print("Usage: python main.py <error_data.csv> [reference_data.csv]")
        sys.exit(1)

    # Load error-induced dataset and reference dataset
    error_data_path = sys.argv[1]
    reference_data_path = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        error_data = pd.read_csv(error_data_path)
        print(f"Loaded error data from {error_data_path}.")
    except Exception as e:
        print(f"Error loading error data file: {e}")
        sys.exit(1)

    try:
        reference_data = pd.read_csv(reference_data_path) if reference_data_path else error_data.copy()
        print(f"Loaded reference data from {reference_data_path or 'error data (default)'}.")
    except Exception as e:
        print(f"Error loading reference data file: {e}")
        sys.exit(1)

    # Detect column types
    print("Detecting column types...")
    numeric_features, categorical_features = detect_column_types(error_data)
    print(f"Numeric Features: {numeric_features}")
    print(f"Categorical Features: {categorical_features}")

    # Initialize Detector and Corrector
    detector = ErrorDetector(reference_data)
    corrector = ErrorCorrector(reference_data)

    # Detect and correct errors
    corrected_data = error_data.copy()

    # Missing values
    print("Detecting and correcting missing values...")
    try:
        corrected_data = corrector.correct_missing_values(corrected_data)
    except Exception as e:
        print(f"Error correcting missing values: {e}")

    # Typographical errors
    print("Detecting and correcting typographical errors...")
    for column in categorical_features:
        try:
            corrected_data = corrector.correct_typographical_errors(corrected_data, column)
        except Exception as e:
            print(f"Error correcting typographical errors in column '{column}': {e}")

    # Out-of-range values
    print("Detecting and correcting out-of-range values...")
    for column in numeric_features:
        try:
            corrected_data = corrector.correct_out_of_range(corrected_data, column)
        except Exception as e:
            print(f"Error correcting out-of-range values in column '{column}': {e}")

    # Save corrected dataset
    output_file = "corrected_data.csv"
    try:
        corrected_data.to_csv(output_file, index=False)
        print(f"Corrected data saved to '{output_file}'.")
    except Exception as e:
        print(f"Error saving corrected data: {e}")

    # Calculate metrics
    calculate_metrics(error_data, corrected_data, reference_data)

    # Generate error-highlighted and corrected-highlighted files
    highlight_errors_and_corrections(
        error_data, corrected_data, "error_highlighted.xlsx", "corrected_highlighted.xlsx"
    )

if __name__ == "__main__":
    main()
